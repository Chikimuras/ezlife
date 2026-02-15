from datetime import datetime, time
from typing import BinaryIO
from uuid import UUID

from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.activity import Activity, Category, Group


class ImportService:
    def __init__(self, session: AsyncSession):
        self.session = session

    async def import_excel(
        self, file: BinaryIO, user_id: UUID
    ) -> dict[str, int | list[str]]:
        logger.info(f"Starting Excel import for user_id={user_id}")
        errors: list[str] = []
        groups_created = 0
        categories_created = 0
        activities_created = 0

        try:
            workbook = load_workbook(filename=file, data_only=True)
            logger.debug(
                f"Excel file loaded successfully, sheets: {workbook.sheetnames}"
            )
        except Exception as e:
            logger.error(f"Failed to load Excel file for user_id={user_id}: {e}")
            return {
                "groups_created": 0,
                "categories_created": 0,
                "activities_created": 0,
                "errors": [f"Failed to load Excel file: {str(e)}"],
            }

        if "Paramètre" in workbook.sheetnames:
            logger.info("Processing 'Paramètre' sheet")
            parametre_sheet = workbook["Paramètre"]
            group_map: dict[str, str] = {}

            for row_idx, row in enumerate(
                parametre_sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                if not row or not row[0]:
                    continue

                try:
                    category_name = row[0]
                    group_name = row[1]
                    priority = row[2] if row[2] is not None else 1
                    min_weekly = row[3] if row[3] is not None else 0.0
                    target_weekly = row[4] if row[4] is not None else 0.0
                    max_weekly = row[5] if row[5] is not None else 0.0
                    unit = row[6] if row[6] else "hours"
                    mandatory_raw = row[7] if len(row) > 7 else ""

                    mandatory = str(mandatory_raw).strip().lower() == "oui"

                    if str(unit).lower() not in ["hours", "minutes", "count"]:
                        unit = "hours"

                    if group_name not in group_map:
                        result = await self.session.execute(
                            select(Group).where(
                                Group.user_id == user_id, Group.name == group_name
                            )
                        )
                        existing_group = result.scalars().first()

                        if existing_group:
                            group_map[group_name] = str(existing_group.id)
                            logger.debug(f"Found existing group: '{group_name}'")
                        else:
                            new_group = Group(
                                user_id=user_id,
                                name=group_name,
                                color=self._get_default_color(len(group_map)),
                            )
                            self.session.add(new_group)
                            await self.session.flush()
                            group_map[group_name] = str(new_group.id)
                            groups_created += 1
                            logger.info(
                                f"Created new group: '{group_name}' (id={new_group.id})"
                            )

                    new_category = Category(
                        user_id=user_id,
                        group_id=group_map[group_name],
                        name=category_name,
                        priority=int(priority),
                        min_weekly_hours=float(min_weekly),
                        target_weekly_hours=float(target_weekly),
                        max_weekly_hours=float(max_weekly),
                        unit=str(unit).lower(),
                        mandatory=mandatory,
                    )
                    self.session.add(new_category)
                    categories_created += 1
                    logger.debug(
                        f"Created category: '{category_name}' in group '{group_name}'"
                    )

                except Exception as e:
                    error_msg = f"Row {row_idx} in Paramètre: {str(e)}"
                    errors.append(error_msg)
                    logger.warning(error_msg)

            logger.success(
                f"'Paramètre' sheet processed: {groups_created} groups, "
                f"{categories_created} categories"
            )

        if "Journal" in workbook.sheetnames:
            logger.info("Processing 'Journal' sheet")
            journal_sheet = workbook["Journal"]

            result = await self.session.execute(
                select(Category).where(Category.user_id == user_id)
            )
            categories = result.scalars().all()
            category_map = {cat.name: str(cat.id) for cat in categories}
            logger.debug(f"Built category map with {len(category_map)} categories")

            for row_idx, row in enumerate(
                journal_sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                if not row or not row[0]:
                    continue

                try:
                    date_value = row[0]
                    start_hour = row[2]
                    start_minute = row[3]
                    end_hour = row[4]
                    end_minute = row[5]
                    category_name = row[9]
                    notes = row[11] if len(row) > 11 and row[11] else None

                    if isinstance(date_value, datetime):
                        activity_date = date_value.date()
                    else:
                        errors.append(f"Row {row_idx} in Journal: Invalid date format")
                        continue

                    if (
                        start_hour is None
                        or start_minute is None
                        or end_hour is None
                        or end_minute is None
                    ):
                        errors.append(f"Row {row_idx} in Journal: Missing time values")
                        continue

                    start_time = time(hour=int(start_hour), minute=int(start_minute))
                    end_time = time(hour=int(end_hour), minute=int(end_minute))

                    if start_time >= end_time:
                        errors.append(
                            f"Row {row_idx} in Journal: "
                            f"End time must be after start time"
                        )
                        continue

                    if category_name not in category_map:
                        errors.append(
                            f"Row {row_idx} in Journal: "
                            f"Category '{category_name}' not found"
                        )
                        continue

                    new_activity = Activity(
                        user_id=user_id,
                        category_id=category_map[category_name],
                        date=activity_date,
                        start_time=start_time,
                        end_time=end_time,
                        notes=notes,
                    )
                    self.session.add(new_activity)
                    activities_created += 1

                except Exception as e:
                    error_msg = f"Row {row_idx} in Journal: {str(e)}"
                    errors.append(error_msg)
                    logger.warning(error_msg)

            logger.success(
                f"'Journal' sheet processed: {activities_created} activities"
            )

        try:
            await self.session.commit()
            logger.success(
                f"Excel import completed for user_id={user_id}: "
                f"{groups_created} groups, {categories_created} categories, "
                f"{activities_created} activities, {len(errors)} errors"
            )
        except Exception as e:
            await self.session.rollback()
            logger.error(f"Database commit failed for user_id={user_id}: {e}")
            return {
                "groups_created": 0,
                "categories_created": 0,
                "activities_created": 0,
                "errors": [f"Database commit failed: {str(e)}"],
            }

        return {
            "groups_created": groups_created,
            "categories_created": categories_created,
            "activities_created": activities_created,
            "errors": errors,
        }

    def _get_default_color(self, index: int) -> str:
        colors = [
            "#FF5733",
            "#33FF57",
            "#3357FF",
            "#FF33F5",
            "#F5FF33",
            "#33F5FF",
            "#FF8C33",
            "#8C33FF",
            "#33FF8C",
            "#FF3333",
        ]
        return colors[index % len(colors)]
